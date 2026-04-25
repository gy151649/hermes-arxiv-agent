#!/usr/bin/env python3
"""
LLM 信息补全脚本：提取作者单位 + 生成中文摘要
处理 pending_llm_ids.txt 中的论文，结果写入 Excel
"""

import os
import json
import re
import sys
from pathlib import Path
from typing import List, Dict, Optional

import fitz  # PyMuPDF
import openpyxl
import openai

# ==================== 配置 ====================
BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "papers_record.xlsx"
PAPERS_DIR = BASE_DIR / "papers"
NEW_PAPERS_JSON = BASE_DIR / "new_papers.json"
PENDING_LLM_IDS_FILE = BASE_DIR / "pending_llm_ids.txt"

# OpenAI 配置（从环境变量读取）
openai.api_key = os.getenv("OPENAI_API_KEY")
openai.base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")
MODEL_NAME = os.getenv("LLM_MODEL", "gpt-4o-mini")

# ==================== 工具函数 ====================

def extract_text_from_pdf_first_two_pages(pdf_path: str) -> str:
    """从 PDF 提取前两页文本"""
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page_num in range(min(2, len(doc))):
            page = doc[page_num]
            text += page.get_text() + "\n"
        doc.close()
        return text
    except Exception as e:
        print(f"[ERROR] Failed to read PDF {pdf_path}: {e}")
        return ""

def extract_affiliations(text: str) -> str:
    """从 PDF 前两页文本提取作者单位，多个用 ; 分隔"""
    # 常见单位提取模式
    # 1. 作者列表后紧跟着的单位信息
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    text_lower = text.lower()
    
    # 尝试多种模式匹配
    affiliations = []
    
    # 模式：寻找作者后括号中的单位，或者作者行下方的单位块
    # 移除摘要前的部分通常包含单位信息
    abstract_pos = text_lower.find("abstract")
    if abstract_pos != -1:
        header_text = text[:abstract_pos]
    else:
        header_text = text
    
    # 正则匹配常见单位格式
    # 匹配 1Department ..., 2University ... 等格式
    dept_pattern = r'(?:^|\W)[0-9]*\s*(Department|Institute|University|College|School|Laboratory|Center|Lab|Google|Microsoft|Meta|IBM|Amazon|Alibaba|Tencent|ByteDance|Huawei)([^,\n]*(?:,[^,\n]*)*)'
    matches = re.findall(dept_pattern, header_text, re.IGNORECASE)
    
    seen = set()
    for match in matches:
        aff = "".join(match).strip()
        if len(aff) > 5 and aff.lower() not in seen:
            affiliations.append(aff)
            seen.add(aff.lower())
    
    # 如果没找到，尝试提取邮箱域名推断机构
    if not affiliations:
        email_pattern = r'[\w.+-]+@([\w-]+\.)+[\w-]+'
        emails = re.findall(email_pattern, header_text)
        for domain in emails:
            if domain:
                aff = domain.strip('.')
                if aff not in seen:
                    affiliations.append(aff)
                    seen.add(aff)
    
    # 清理结果
    cleaned = []
    for aff in affiliations:
        # 去掉数字前缀
        aff_clean = re.sub(r'^[0-9\s\*,]+', '', aff).strip()
        if len(aff_clean) > 3 and aff_clean not in cleaned:
            cleaned.append(aff_clean)
    
    # 多个用分号分隔
    result = "; ".join(cleaned)
    # 如果太长截断
    if len(result) > 500:
        result = result[:497] + "..."
    return result

def generate_chinese_summary(abstract: str, title: str) -> str:
    """基于英文摘要生成 90-150 字中文摘要"""
    if not openai.api_key:
        return "[LLM 未配置 需要手动补全]"
    
    prompt = f"""请基于以下 arXiv 论文的标题和英文摘要，生成一篇 90-150 字的中文摘要。
要求：
1. 概括研究背景、核心方法、主要结论
2. 语言简洁流畅，符合中文学术表达习惯
3. 严格控制字数在 90-150 字之间
4. 不要添加原文没有的信息

标题：{title}

英文摘要：
{abstract}

请直接输出中文摘要，不要其他内容：
"""
    
    try:
        response = openai.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": "你是一个专业的 AI 学术论文翻译总结助手。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=300,
        )
        summary = response.choices[0].message.content.strip()
        # 清理可能的引号
        summary = summary.strip('"\'')
        return summary
    except Exception as e:
        print(f"[ERROR] LLM summary generation failed: {e}")
        return "[LLM 调用失败 需要手动重试]"

def load_pending_papers() -> List[Dict]:
    """从 new_papers.json 加载待处理论文"""
    with open(NEW_PAPERS_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("papers_to_process", data.get("new_papers", []))

def update_excel_affiliations_and_summary(arxiv_id: str, affiliations: str, summary_cn: str) -> bool:
    """更新 Excel 中对应行的 affiliations 和 summary_cn"""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Papers" not in wb.sheetnames:
            print(f"[ERROR] Sheet 'Papers' not found")
            return False
        
        ws = wb["Papers"]
        # 构建索引
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_index = {str(v): i + 1 for i, v in enumerate(header_row) if v is not None}
        
        arxiv_col = header_index.get("arxiv_id")
        aff_col = header_index.get("affiliations")
        summary_cn_col = header_index.get("summary_cn")
        
        if not all([arxiv_col, aff_col, summary_cn_col]):
            print(f"[ERROR] Required columns not found")
            return False
        
        # 查找对应行
        found = False
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=arxiv_col).value
            if cell_val and str(cell_val).strip() == arxiv_id:
                ws.cell(row=r, column=aff_col, value=affiliations)
                ws.cell(row=r, column=summary_cn_col, value=summary_cn)
                found = True
                break
        
        if found:
            wb.save(EXCEL_FILE)
            print(f"[INFO] Excel updated: {arxiv_id}")
            return True
        else:
            print(f"[WARN] arxiv_id {arxiv_id} not found in Excel")
            return False
    except Exception as e:
        print(f"[ERROR] Failed to update Excel: {e}")
        return False

def main():
    print("=" * 60)
    print(f"[START] LLM Info Completion | Path: {BASE_DIR}")
    print(f"[INFO] Using model: {MODEL_NAME}")
    
    # 检查环境
    if not openai.api_key:
        print("[WARN] OPENAI_API_KEY not set, LLM summary will be placeholder")
    
    # 加载待处理论文
    pending = load_pending_papers()
    print(f"[INFO] Found {len(pending)} papers to process")
    
    success_count = 0
    fail_count = 0
    
    for i, paper in enumerate(pending, 1):
        arxiv_id = paper["arxiv_id"]
        title = paper["title"]
        abstract = paper["abstract"]
        pdf_filename = paper.get("pdf_filename", f"{arxiv_id}.pdf")
        pdf_path = PAPERS_DIR / pdf_filename
        
        print(f"\n[{i}/{len(pending)}] Processing: {arxiv_id} - {title[:50]}...")
        
        if not pdf_path.exists():
            print(f"[ERROR] PDF not found: {pdf_path}")
            fail_count += 1
            continue
        
        # 1. 提取作者单位
        pdf_text = extract_text_from_pdf_first_two_pages(str(pdf_path))
        affiliations = extract_affiliations(pdf_text)
        print(f"[INFO] Extracted affiliations: {affiliations[:100]}...")
        
        # 2. 生成中文摘要
        summary_cn = generate_chinese_summary(abstract, title)
        print(f"[INFO] Generated summary_cn ({len(summary_cn)} chars): {summary_cn}")
        
        # 3. 更新 Excel
        if update_excel_affiliations_and_summary(arxiv_id, affiliations, summary_cn):
            success_count += 1
        else:
            fail_count += 1
    
    print("\n" + "=" * 60)
    print(f"[DONE] Total: {len(pending)} | Success: {success_count} | Failed: {fail_count}")
    print("=" * 60)

if __name__ == "__main__":
    main()
